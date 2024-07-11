# -*- coding: utf-8 -*-

import time
import os

# import io
import shelve
import json

# import threading

from threading import Timer
from htmlc import HTML
from operator import itemgetter, attrgetter
from tkinter import *
from tkinter.filedialog import Open, SaveAs, asksaveasfilename
from tkinter.messagebox import askyesno
from tkinter.simpledialog import askstring, askinteger, askfloat
import tkinter as tk

try:
    import paho.mqtt.client as paho

    mqttsupport = True
except ImportError:
    mqttsupport = False
    print("paho.mqtt.client doesn't exist, please install it.")

import os
import socket
import ssl
from time import sleep
from random import uniform

# from tkinter import ttk

from ckilpailija import kilpailija

try:
    from openpyxl import Workbook

    xlsxsupport = True
except ImportError:
    xlsxsupport = False
    print("Openpyxl doesn't exist, please install it.")

import csv


""" Ohje sorttaamiseen http://wiki.python.org/moin/HowTo/Sorting """
""" http://www.feig-electronics.com/uploads/media/Data_sheet_LRU3000_3500_01.pdf """

""" http://www.daniweb.com/software-development/python/threads/299142/array-of-my-own-classes """


connflag = False


def on_connect(client, userdata, flags, rc):
    global connflag
    connflag = True
    print("Connection returned result: " + str(rc))


def on_message(client, userdata, msg):
    print(msg.topic + " " + str(msg.payload))


class Kilpailu(Frame):
    luokkafound = 0
    startfiledir = "."
    ftypes = [("CSV", ".csv")]
    ftypes2 = [("DAT", ".dat")]
    """ Implements a stop watch frame widget. """

    def __init__(self, parent=None, **kw):
        self.__startTime = 0
        self.__endTime = 0
        self.__competition_running = False
        self.__kilpailijoita = 0
        self._competitors_finished = 0
        self.__competiontName = "lobby"
        self.__matka = []
        self.temppi = []
        self.competitors = []
        self.found = 0
        Frame.__init__(self, parent, kw)
        # Frame.__init__(self, parent, kw)
        """self.pack(expand=YES,fill=BOTH)"""
        self._start = 0.0
        self._elapsedtime = 0.0
        self._pausestart = 0.0
        self._elapsedpause = 0.0
        self._running = 0
        self.timestr = StringVar()
        self.strfinished = StringVar()
        self.syotanumero = StringVar()
        self.competitionphase = StringVar()
        self.makeWidgets()
        self.temporary = []
        self.temporary2 = []
        self.luokat = []
        self.vaihelaskuri = [0, 0, 0, 0, 0, 0, 0]
        self._timeamount = 1
        self.timeupdateinterval = Timer(0.01, self._update)

    def lueosallistujat(self):
        self.openDialog = Open(initialdir=self.startfiledir, filetypes=self.ftypes)
        file = self.openDialog.show()
        self.vaihelaskuri[0] = 0
        self.vaihelaskuri[1] = 0
        self.vaihelaskuri[2] = 0
        self.vaihelaskuri[3] = 0
        if not file:
            return

        if not os.path.isfile(file):
            self.writeToLog("Ajanotto by Juha Viitanen 2016 Tiedostoa ei voitu avata " + file)

        try:
            self.dbase = shelve.open(
                "backlog/backlog" + time.strftime("%d_%m_%y_%H_%M_%S", time.localtime(time.time())) + "shv"
            )
            with open(file, newline="", encoding="utf-8") as csvfile:
                spamreader = csv.reader(csvfile, delimiter=";", quotechar="|")
                """print(spamreader)"""
                for row in spamreader:
                    self.competitors.append(kilpailija(row[2], row[1], "+358........", row[3], row[4], row[0]))
                    self.__kilpailijoita = self.__kilpailijoita + 1
                    for luokka in self.luokat:
                        if luokka == row[4]:
                            self.luokkafound = 1
                            break
                    if self.luokkafound == 0:
                        self.luokat.append(row[4])
                    else:
                        self.luokkafound = 0

                    # self.writeToLog(', '.join(row))
                # self.writeToLog(self.luokat)
                self.writeToLog("Ladattiin " + str(self.__kilpailijoita) + " kilpailijaa järjestelmään")

                for obj in self.competitors:
                    self.dbase[obj.bibnumber] = obj
                self.dbase.close()
                """self.writeToLog(('%s'%self.__kilpailijoita))"""
        except:
            self.writeToLog("CSV-tiedostoa ei löytynyt tai sitä ei voi avatata.")

    def lueosallistujatshelve(self):
        self.openDialog = Open(initialdir=self.startfiledir, filetypes=self.ftypes2)
        file = self.openDialog.show()

        if not file:
            print("plaa")
            return

        if not os.path.isfile(file):
            self.writeToLog("Ajanotto by Juha Viitanen Tiedostoa ei voitu avata " + file)

        try:
            print("tuliko" + os.path.splitext(file)[0])
            self.dbase = shelve.open(os.path.splitext(file)[0])
            # print(list(self.dbase.keys()))
            for key in list(self.dbase.keys()):
                # print(self.dbase[key].bibnumber)
                self.competitors.append(self.dbase[key])
                self.__kilpailijoita = self.__kilpailijoita + 1
                for luokka in self.luokat:
                    if luokka == self.dbase[key].kilpasarja:
                        self.luokkafound = 1
                        break
                if self.luokkafound == 0:
                    self.luokat.append(self.dbase[key].kilpasarja)
                else:
                    self.luokkafound = 0
            print(list(self.luokat))
            self.writeToLog("Ladattiin " + str(self.__kilpailijoita) + " kilpailijaa järjestelmään")
        except:
            print("Jotain meni pieleen")

    def makeWidgets(self):
        """Make the time label."""
        self.makemenu()
        l = Label(self, textvariable=self.timestr, font=("arial", 16, "bold"), bg="black", fg="yellow")
        self.dnfbutton = Button(self, text="DNF", command=self.didnotfinish, font=("arial", 16, "bold"))
        self.dnsbutton = Button(self, text="DNS", command=self.didnotstart, font=("arial", 16, "bold"))
        self.dsqbutton = Button(self, text="DSQ", command=self.disqualify, font=("arial", 16, "bold"))
        self.startbutton = Button(self, text="START", command=self.aloitakilpailu, font=("arial", 16, "bold"))

        self.strfinished.set("")
        self.competitionphase.set("Phase")
        l.config(font=("arial", 24, "bold"))
        self.timestr.set("Kilpailu ei ole käynnissä")
        l.pack()

        """ log = Text(self,relief=SUNKEN).grid(row=2) """
        """ self.log = log """

        # Button(self, text='Stop', command=self.lopetakilpailu,font=('arial',16,'bold')).pack(side=RIGHT, expand=YES, fill=BOTH)
        if xlsxsupport == True:
            Button(self, text="XLSX", command=self.writetoxlsx, font=("arial", 16, "bold")).pack(
                side=RIGHT, expand=YES, fill=BOTH
            )
        # Button(self, text='TXT', command=self.writeOfficialTimes,font=('arial',16,'bold')).pack(side=RIGHT, expand=YES, fill=BOTH)
        Button(self, text="HTML", command=self.writeHTML, font=("arial", 16, "bold")).pack(
            side=RIGHT, expand=YES, fill=BOTH
        )
        # Button(self, text='Startlist', command=self.writetotxt,font=('arial',16,'bold')).pack(side=BOTTOM, expand=YES, fill=BOTH)

        self.scroll = Scrollbar(self)
        self.log = Text(
            self,
            state="disabled",
            width=120,
            height=30,
            wrap="none",
            bg="black",
            fg="yellow",
            font=("Courier", 10, "bold"),
        )
        finished = Label(self, textvariable=self.strfinished)
        finished.config(font=("arial", 16, "bold"))

        phase = Label(self, textvariable=self.competitionphase)
        phase.config(font=("arial", 16, "bold"))

        self.numero = Entry(
            self,
            textvariable=self.syotanumero,
            font=("arial", 40, "bold"),
            width=5,
            justify="center",
            bg="yellow",
            fg="black",
        )

        self.scroll.config(command=self.log.yview)
        self.scroll.pack(side=RIGHT, expand=YES, fill=BOTH)
        self.log.pack()
        self.log.config(yscrollcommand=self.scroll.set)
        self.numero.pack()
        finished.pack()
        phase.pack()

        self.dnfbutton.pack(side=LEFT, expand=NO, fill=BOTH)
        self.dnsbutton.pack(side=LEFT, expand=NO, fill=BOTH)
        self.dsqbutton.pack(side=LEFT, expand=NO, fill=BOTH)
        self.startbutton.pack(side=LEFT, expand=YES, fill=BOTH)
        """self.scroll.pack(side=RIGHT,fill=Y)"""
        """slf.log.pack(side=LEFT,expand=YES, fill=BOTH)"""

    def makemenu(self):
        self.menubar = Menu(self.master)
        self.master.config(menu=self.menubar)
        file = Menu(self.menubar)
        file.add_command(label="TXT virallinen", command=lambda: self.writeOfficialTimes(0, False))
        file.add_command(label="TXT kilpanumerolla", command=lambda: self.writeOfficialTimes(1, False))
        file.add_command(label="HTML", command=self.writeHTML)
        if xlsxsupport == True:
            file.add_command(label="XLSX")
        self.menubar.add_cascade(label="Vie tiedot", menu=file)

        kilpailu = Menu(self.master)
        kilpailu.add_command(label="Kilpailun tiedot", command=self.syotatiedot)
        kilpailu.add_command(label="Start", command=self.aloitakilpailu)
        kilpailu.add_command(label="Stop", command=self.lopetakilpailu)
        self.menubar.add_cascade(label="Kilpailu", menu=kilpailu)

        kilpailija = Menu(self.master)
        kilpailija.add_command(label="Clear competitor", command=self.clearcompetitor)
        self.menubar.add_cascade(label="kilpailija", menu=kilpailija)

        tiedot = Menu(self.master)
        tiedot.add_command(label="Lue kilpanauhoite .dat", command=self.lueosallistujatshelve)
        tiedot.add_command(label="Lue kilpailijat .csv", command=self.lueosallistujat)
        self.menubar.add_cascade(label="Datat", menu=tiedot)

        tallenna = Menu(self.master)
        if xlsxsupport == True:
            tallenna.add_command(label="XLSX", command=self.lueosallistujatshelve)
        tallenna.add_command(label="TXT virallinen", command=lambda: self.writeOfficialTimes(0, True))
        tallenna.add_command(label="TXT kilpanumerolla", command=lambda: self.writeOfficialTimes(1, True))
        tallenna.add_command(label="dat-tiedosto", command=self.lueosallistujatshelve)
        tallenna.add_command(label="JSON", command=self.lueosallistujatshelve)
        self.menubar.add_cascade(label="Tallenna", menu=tallenna)

    def _update(self):
        self.timestr.set(
            time.strftime("%H:%M:%S", time.gmtime(self.getKisaaika()))
            + (",%02d" % int(((self.getKisaaika() - int(self.getKisaaika())) * 100)))
        )
        self._timer = self.after(100, self._update)

    def _setTime(self, elap):
        print("update request")
        self.timestr.set(time.strftime("%H:%M:%S", time.gmtime(self.getKisaaika())))

    def status(self):
        return self.__competition_running

    def aloitusaika(self):
        return self.__startTime

    def lopetusaika(self):
        return self.__endTime

    def syotatiedot(self):
        #         i = 0
        self.__competiontName = askstring("Kilpailun nimi", "Anna kilpailun nimi")  # input("Kilpailun nimi: ")
        self._timeamount = askinteger("Väliaikojen määrä", "Anna väliaikojen määrä")

    def clearcompetitor(self):
        #         i = 0
        bibnum = askstring("Kilpailijan numero", "Kilpailijan numero")  # input("Kilpailun nimi: ")
        self.clearcompetitorhelper(bibnum)
        # askinteger
        """while i < self._timeamount:
            self.__matka.append(askfloat('Matka: '+ str(i+1),'Vaiheen: '+ str(i+1)+' matka'))
            i +=1"""

    def getKisaaika(self):
        return time.time() - self.__startTime

    def aloitakilpailu(self):
        if self.__competition_running == False:
            self.__startTime = time.time()
            self.__competition_running = True
            self.startbutton["state"] = tk.DISABLED
            # print("Kilpailu käynnistetty: ",time.strftime("%H:%M:%S",time.localtime(self.__startTime)))
            self.writeToLog(("Kilpailu käynnistetty: " + time.strftime("%H:%M:%S", time.localtime(self.__startTime))))
            self.writeCompetitionTimes(
                (
                    ("<h1>Kilpailu käynnistetty: " + time.strftime("%H:%M:%S", time.localtime(self.__startTime)))
                    + "</h1>"
                ).replace("ä", "&auml")
            )
            self._update()
            # self.timeupdateinterval.start()
            # f.write("Kilpailu käynnistetty: ")
            # f.write(time.strftime("%H:%M:%S",time.localtime(self.__startTime))+"\n")
            # f.close()
        else:
            # print("Kilpailu on jo käynnissä!")
            self.writeToLog("Kilpailu on jo käynnissä!")

    def lopetakilpailu(self):
        # return
        if self.__competition_running == True:
            ans = askyesno("Veryfy exit", "Oletko varma, että haluat pysäyttää ajastimen?")
        else:
            return
        if ans == False:
            return
        try:
            self.dbase.close()
        except:
            print("Ei määritettyä tietokantaa")

        if self.__competition_running == True:
            self.startbutton["state"] = tk.NORMAL
            self.__endTime = time.time()
            self.__competition_running = False
            # print("Kilpailu pysäytetty: "+time.strftime("%H:%M:%S",time.localtime(self.__endTime)))
            self.writeToLog(("Kilpailu pysäytetty: " + time.strftime("%H:%M:%S", time.localtime(self.__endTime))))
            kesto = self.__endTime - self.__startTime
            # print("Kilpailu kesti: ", time.strftime("%H:%M:%S",time.gmtime(kesto)))
            self.writeToLog(("Kilpailu kesti: " + time.strftime("%H:%M:%S", time.gmtime(kesto))))

            self.timestr.set(time.strftime("%H:%M:%S", time.gmtime(kesto)))
            self.after_cancel(self._timer)

            # print("Kilpailu kesti ",kesto/60.0/60," tuntia ",kesto/60.0," minuuttia ja ",kesto," sekuntia")
        else:
            self.writeToLog("Kilpailu on jo pysäytetty")
        print(json.dumps(self.luokat, default=lambda o: o.__dict__, sort_keys=True, indent=4, ensure_ascii=True))

    def writeToLog(self, msg):
        self.numlines = self.log.index("end - 1 line").split(".")[0]
        self.log["state"] = "normal"
        if self.numlines == 24:
            self.log.delete("1.0", "2.0")
        if self.log.index("end-1c") != "1.0":
            self.log.insert("1.0", "\n")
        self.log.insert("1.0", msg)
        self.log["state"] = "disabled"

    def writeCompetitionTimes(self, msg):
        f = open("export.txt", "a")
        f.write(msg + "\n")
        f.close()

    def writeToLogXLSX(self, msg):
        self.numlines = self.log.index("end - 1 line").split(".")[0]
        self.log["state"] = "normal"
        if self.numlines == 24:
            self.log.delete("1.0", "2.0")
        if self.log.index("end-1c") != "1.0":
            self.log.insert("1.0", "\n")
        self.log.insert("1.0", msg)
        self.log["state"] = "disabled"

    def ConvertTimeToString(self, aika):
        return time.strftime("%H:%M:%S", time.gmtime(aika))

    def ConvertTimeToStringAccurate(self, aika):
        return time.strftime("%H:%M:%S", time.gmtime(aika)) + (",%02d" % int(((aika - int(aika)) * 100)))

    def didnotfinish(self):
        if len(self.syotanumero.get()) > 0:
            for obj in self.competitors:
                """print(obj.bibnumber)"""
                if obj.bibnumber == self.syotanumero.get():
                    """print('sama numero jo kirjattu')"""
                    if obj.isStatusOn() == True:
                        self._competitors_finished = self._competitors_finished - 1

                    if obj.isStatusOn() == False:
                        self._competitors_finished = self._competitors_finished + 1
                    obj.DNF()
                    # self.strfinished.set(("#%s"%self._competitors_finished))
                    # self.writeToLog(("%s" % self._competitors_finished) + "\t" + self.syotanumero.get() + "\t" + "DNF" + "\t" + obj.etunimi + "\t" + obj.sukunimi + "\t" + obj.seura + "\t" + obj.kilpasarja)
                    self.getpositionetc(obj.bibnumber)
                    self.syotanumero.set("")
                    break

    def clearcompetitorhelper(self, bibnum):
        for obj in self.competitors:
            if obj.bibnumber == bibnum:
                if obj.GetTimeAmount() > 0:
                    self._competitors_finished = self._competitors_finished - 1
                    contents = str("(" + obj.bibnumber + ") ")
                    contents = str(contents + obj.etunimi + " ")
                    contents = str(contents + obj.sukunimi + " ")
                    contents = str(contents + obj.seura + " CLEARED")
                    self.writeToLog(contents)
                    obj.clear()
                break

    def didnotstart(self):
        if len(self.syotanumero.get()) > 0:
            for obj in self.competitors:
                """print(obj.bibnumber)"""
                if obj.bibnumber == self.syotanumero.get():
                    """print('sama numero jo kirjattu')"""
                    if obj.isStatusOn() == True:
                        self._competitors_finished = self._competitors_finished - 1

                    if obj.isStatusOn() == False:
                        self._competitors_finished = self._competitors_finished + 1
                    obj.DNS()
                    # self.strfinished.set(("#%s"%self._competitors_finished))
                    # self.writeToLog(("%s" % self._competitors_finished) + "\t" + self.syotanumero.get() + "\t" + "DNS" + "\t" + obj.etunimi + "\t" + obj.sukunimi + "\t" + obj.seura + "\t" + obj.kilpasarja)
                    self.getpositionetc(obj.bibnumber)
                    self.syotanumero.set("")
                    break

    def disqualify(self):
        if len(self.syotanumero.get()) > 0:
            for obj in self.competitors:
                """print(obj.bibnumber)"""
                if obj.bibnumber == self.syotanumero.get():
                    """print('sama numero jo kirjattu')"""
                    if obj.isStatusOn() == True:
                        self._competitors_finished = self._competitors_finished - 1

                    if obj.isStatusOn() == False:
                        self._competitors_finished = self._competitors_finished + 1
                    obj.DISQUALIFY()
                    # self.strfinished.set(("#%s"%self._competitors_finished))
                    # self.writeToLog(("%s" % self._competitors_finished) + "\t" + self.syotanumero.get() + "\t" + "DNS" + "\t" + obj.etunimi + "\t" + obj.sukunimi + "\t" + obj.seura + "\t" + obj.kilpasarja)
                    self.getpositionetc(obj.bibnumber)
                    self.syotanumero.set("")
                    break

    def getpositionetc(self, competitionnumber):
        contents = ""
        retobj = 0
        self.temporary2 = sorted(self.competitors, key=attrgetter("totaltime"))
        self.temporary = sorted(self.temporary2, key=attrgetter("valiaikamaara"), reverse=True)

        for luokka in self.luokat:
            position = 0
            # contents = str(contents + luokka+"\n")
            for obj in self.temporary:
                if luokka == obj.kilpasarja:
                    if obj.totaltime != 9999999999:
                        position = position + 1
                    if position == 1:
                        nr1pos = obj
                    if competitionnumber == obj.bibnumber:
                        columniterator = 0
                        if obj.totaltime != 9999999999:
                            contents = str(
                                contents
                                + "("
                                + str(self._competitors_finished)
                                + "/"
                                + str(self.__kilpailijoita)
                                + ") "
                            )
                            contents = str(contents + " ")
                        else:
                            contents = str(
                                contents
                                + "("
                                + str(self._competitors_finished)
                                + "/"
                                + str(self.__kilpailijoita)
                                + ") "
                            )
                            contents = str(contents + " ")

                        contents += str("(" + obj.bibnumber + ") ")
                        contents = str(contents + obj.etunimi + " ")
                        contents = str(contents + obj.sukunimi + " ")
                        contents = str(contents + obj.seura + " |")
                        if obj.isStatusOn() == False:
                            contents = (contents + "sij. " + str(position) + ". " + luokka + " ") + "|"
                        else:
                            contents = contents + obj.getStatus()
                        obj.Sijoitus(position)
                        retobj = obj
                        #                     f.write(obj.bibnumber+" ")
                        if obj.lasttime != 0:
                            # if(position == 1):
                            #    nr1pos = obj;
                            if obj.dnf == False and obj.dns == False and obj.dsq == False:
                                if obj.GetTimeAmount() > 1:
                                    if position == 1:
                                        contents = str(contents + self.ConvertTimeToString(obj.totaltime) + " ")
                                        self.writeCompetitionTimes(
                                            (
                                                "<dt><strong>"
                                                + str(obj.bibnumber)
                                                + " "
                                                + obj.etunimi
                                                + " "
                                                + obj.sukunimi
                                                + "</strong></dt><dd>"
                                                + luokka
                                                + "</dd>"
                                                + "<dd>Sija: "
                                                + str(position)
                                                + "</dd><dd>"
                                                + obj.seura
                                                + "</dd><dd>"
                                                + self.ConvertTimeToStringAccurate(obj.totaltime)
                                                + "</dd>"
                                            ).replace("ä", "&auml")
                                        )
                                    else:
                                        contents = str(
                                            contents
                                            + self.ConvertTimeToString(obj.totaltime)
                                            + " +"
                                            + self.ConvertTimeToString(obj.totaltime - nr1pos.totaltime)
                                            + " "
                                        )
                                        self.writeCompetitionTimes(
                                            (
                                                "<dt><strong>"
                                                + str(obj.bibnumber)
                                                + " "
                                                + obj.etunimi
                                                + " "
                                                + obj.sukunimi
                                                + "</strong></dt><dd>"
                                                + luokka
                                                + "</dd>"
                                                + "<dd>Sija: "
                                                + str(position)
                                                + "</dd><dd>"
                                                + obj.seura
                                                + "</dd><dd>"
                                                + self.ConvertTimeToStringAccurate(obj.totaltime)
                                                + " +"
                                                + self.ConvertTimeToString(obj.totaltime - nr1pos.totaltime)
                                                + "</dd>"
                                            ).replace("ä", "&auml")
                                        )
                                    for aika in obj.valiajat:
                                        contents = str(contents + self.ConvertTimeToString(aika) + " ")
                                        columniterator = columniterator + 1
                                    if len(obj.valiajat) == 1:
                                        contents = str(contents + " ")
                                        contents = str(contents + " ")
                                    if len(obj.valiajat) == 2:
                                        contents = str(contents + " ")

                                    columniterator = 0
                                else:
                                    if position == 1:
                                        contents = str(contents + self.ConvertTimeToString(obj.totaltime) + " ")
                                        self.writeCompetitionTimes(
                                            (
                                                "<dt><strong>"
                                                + str(obj.bibnumber)
                                                + " "
                                                + obj.etunimi
                                                + " "
                                                + obj.sukunimi
                                                + "</strong></dt><dd>"
                                                + luokka
                                                + "</dd>"
                                                + "<dd>Sija: "
                                                + str(position)
                                                + "</dd><dd>"
                                                + obj.seura
                                                + "</dd><dd>"
                                                + self.ConvertTimeToStringAccurate(obj.totaltime)
                                                + "</dd>"
                                            ).replace("ä", "&auml")
                                        )
                                    else:
                                        contents = str(
                                            contents
                                            + self.ConvertTimeToString(obj.totaltime)
                                            + " +"
                                            + self.ConvertTimeToString(obj.totaltime - nr1pos.totaltime)
                                            + " "
                                        )
                                        self.writeCompetitionTimes(
                                            (
                                                "<dt><strong>"
                                                + str(obj.bibnumber)
                                                + " "
                                                + obj.etunimi
                                                + " "
                                                + obj.sukunimi
                                                + "</strong></dt><dd>"
                                                + luokka
                                                + "</dd>"
                                                + "<dd>Sija: "
                                                + str(position)
                                                + "</dd><dd>"
                                                + obj.seura
                                                + "</dd><dd>"
                                                + self.ConvertTimeToStringAccurate(obj.totaltime)
                                                + " +"
                                                + self.ConvertTimeToString(obj.totaltime - nr1pos.totaltime)
                                                + "</dd>"
                                            ).replace("ä", "&auml")
                                        )
                                        # print("+" + self.ConvertTimeToString(obj.totaltime - nr1pos.totaltime))
                                    contents = str(contents + " ")
                                    contents = str(contents + " ")
                                    contents = str(contents + " ")
                            else:
                                contents = str(contents + obj.getStatus())
                        else:
                            if obj.dnf == False and obj.dns == False and obj.dsq == False:
                                contents = str(contents + obj.getStatus())
                            else:
                                contents = str(contents + " ")
                                contents = str(contents + " ")
                                contents = str(contents + " ")
                                contents = str(contents + " ")

                        contents = str(contents)

        # print(contents)
        self.writeToLog(contents)
        if mqttsupport == True:
            if connflag == True:
                mqttc.publish("competitions/" + self.__competiontName, retobj.toJSON(), qos=1)
                print(retobj)
                # print(json.dumps(retobj))
                print(retobj.toJSON())
                """print(contents)"""
            else:
                print("waiting for connection...")
        contents = "0"
        return contents

    def setfocus(self):
        """print(len(self.syotanumero.get()))
        print(self.numero.focus_set())"""
        self.numero.focus_set()

        if self.__competition_running == True:
            if len(self.syotanumero.get()) > 0:
                self.found = 0
                for obj in self.competitors:
                    """print(obj.bibnumber)"""
                    if obj.bibnumber == self.syotanumero.get():
                        # print("Tasta tuli")
                        """print('sama numero jo kirjattu')"""
                        try:
                            obj.kirjaaAika(self.getKisaaika(), self._timeamount)
                            self.found = 1
                            """self.vaihelaskuri[obj.GetTimeAmount()] = self.vaihelaskuri[obj.GetTimeAmount()]+1"""
                            """self.competitionphase.set("1:"+str(self.vaihelaskuri[1])+"2:"+str(self.vaihelaskuri[2])+"3"+str(self.vaihelaskuri[3]))"""
                        except:
                            print("ei numeroa")
                        break
                    else:
                        self.found = 0
                    """elif obj.bibnumber != self.syotanumero.get():
                        kilpatemp = kilpailija('Etunimi','Sukunimi','Puhelinnumero','Seura','M',self.syotanumero.get())
                        kilpatemp.kirjaaAika(time.strftime("%H:%M:%S",time.gmtime(self.getKisaaika())))
                        self.competitors.append(kilpatemp)
                        print('kilpailijaa ei ole, lisätään')
                        print(len(self.competitors))
                        break"""
                if self.found == 0:
                    self.kilpatemp = kilpailija(
                        "Etunimi", "Sukunimi", "-", "-", "Tarkistettavat:", self.syotanumero.get()
                    )
                    self.__kilpailijoita = self.__kilpailijoita + 1
                    for luokka in self.luokat:
                        if luokka == "Tarkistettavat:":
                            self.luokkafound = 1
                            break
                    if self.luokkafound == 0:
                        self.luokat.append("Tarkistettavat:")
                    else:
                        self.luokkafound = 0

                    self.kilpatemp.kirjaaAika(self.getKisaaika(), self._timeamount)
                    obj = self.kilpatemp
                    self.competitors.append(self.kilpatemp)

                if obj.GetTimeAmount() > 0:
                    self._competitors_finished = self._competitors_finished + 1
                # self.writeToLog(self.syotanumero.get()+"\t"+self.ConvertTimeToStringAccurate(obj.totaltime)+"\t"+obj.etunimi+"\t"+obj.sukunimi+"\t"+obj.seura+"\t"+obj.kilpasarja)
                # self.writeCompetitionTimes(self.syotanumero.get()+"\t"+self.ConvertTimeToString(obj.totaltime)+"\t"+self.ConvertTimeToStringAccurate(obj.totaltime)+"\t"+obj.etunimi+"\t"+obj.sukunimi+"\t"+obj.seura+"\t"+obj.kilpasarja)
                self.strfinished.set(("#%s" % self._competitors_finished))
                self.getpositionetc(self.syotanumero.get())

                try:
                    self.dbase = shelve.open("backlog/backlog")
                    self.dbase[self.syotanumero.get()] = obj
                    print("db update")
                except:
                    print("Tietokannanlatausvirhe")
                # self.writeCompetitionTimes(("<dt><strong>"+self.syotanumero.get()+" "+obj.etunimi+" "+obj.sukunimi+"</strong></dt><dd>"+obj.kilpasarja+"</dd><dd>"+obj.seura+"</dd><dd>"+self.ConvertTimeToStringAccurate(obj.totaltime)+"</dd>").replace("ä","&auml"))
                self.syotanumero.set("")
                self.strfinished.set("")
                # self.
                """self.competitionphase.set('narf')"""

                self.writeHTML()
                """ self.competitors.append(kilpailija('Etunimi','Sukunimi','Puhelinnumero','Seura','Sarja','M',time.strftime("%H:%M:%S",time.gmtime(self.getKisaaika()))))"""

                """ for obj in self.competitors:
                    print(obj)"""
        else:
            self.writeToLog("Kilpailu ei ole käynnissä")

    def addline(self):
        self.writeToLog("-------------------------------------")
        # print(self.log.get(END))

    def getname(self):
        kirjauslkm = 0
        if len(self.syotanumero.get()) > 0:
            self.found = 0
            for obj in self.competitors:
                """print(obj.bibnumber)"""
                if obj.bibnumber == self.syotanumero.get():
                    kirjauslkm = len(obj.ajat) + 1
                    if kirjauslkm == self._timeamount:
                        self.strfinished.set(
                            obj.getStatus()
                            + "+"
                            + ("MAALI! V.aika %d:" % kirjauslkm)
                            + obj.etunimi
                            + " "
                            + obj.sukunimi
                            + ", "
                            + obj.seura
                            + ", "
                            + obj.kilpasarja
                        )
                    else:
                        self.strfinished.set(
                            obj.getStatus()
                            + " "
                            + ("V.aika %d:" % kirjauslkm)
                            + obj.etunimi
                            + " "
                            + obj.sukunimi
                            + ", "
                            + obj.seura
                            + ", "
                            + obj.kilpasarja
                        )
                    break
                else:
                    self.strfinished.set("Numeroa ei ole")
        else:
            self.strfinished.set("")

    def writeHTML(self):
        # return 0
        h = HTML()
        h.h1(self.__competiontName)

        h.p("Ajat")
        t = h.table(border="1", width="100%")
        #         for i in range(2):
        r = t.tr()
        r.td("")
        r.td("")
        r.td("")
        r.td("")
        r.td("#")
        r.td("Total")
        r.td("Time1")
        r.td("Time2")
        r.td("Time3")

        self.temporary2 = sorted(self.competitors, key=attrgetter("totaltime"))
        self.temporary = sorted(self.temporary2, key=attrgetter("valiaikamaara"), reverse=True)

        for luokka in self.luokat:
            r = t.tr
            r.td(luokka, colspan="9", bgcolor="grey", style="color:white")

            position = 0
            for obj in self.temporary:
                if luokka == obj.kilpasarja:
                    r = t.tr
                    columniterator = 0
                    if obj.totaltime != 9999999999:
                        # print(obj.totaltime)
                        position = position + 1
                        r.td(str(position))
                    else:
                        r.td("-")
                    r.td(obj.etunimi)
                    r.td(obj.sukunimi)
                    r.td(obj.seura)
                    r.td(obj.bibnumber)
                    if obj.lasttime != 0:
                        if position == 1:
                            nr1pos = obj
                        if obj.dnf == False and obj.dns == False and obj.dsq == False:
                            if obj.GetTimeAmount() > 1:
                                if position == 1:
                                    r.td(self.ConvertTimeToString(obj.totaltime))
                                else:
                                    r.td(
                                        self.ConvertTimeToString(obj.totaltime)
                                        + " +"
                                        + self.ConvertTimeToString(obj.totaltime - nr1pos.totaltime)
                                    )
                                for aika in obj.valiajat:
                                    r.td(self.ConvertTimeToString(aika))
                                    columniterator = columniterator + 1
                                if len(obj.valiajat) == 1:
                                    r.td("")
                                    r.td("")
                                if len(obj.valiajat) == 2:
                                    r.td("")
                                columniterator = 0
                            else:
                                if position == 1:
                                    r.td(self.ConvertTimeToString(obj.totaltime))
                                else:
                                    r.td(
                                        self.ConvertTimeToString(obj.totaltime)
                                        + " +"
                                        + self.ConvertTimeToString(obj.totaltime - nr1pos.totaltime)
                                    )
                                    # print("+" + self.ConvertTimeToString(obj.totaltime - nr1pos.totaltime))
                                r.td("")
                                r.td("")
                                r.td("")
                        else:
                            r.td("DNF")
                    else:
                        if obj.dnf == False and obj.dns == False and obj.dsq == False:
                            r.td("DNF")
                        else:
                            r.td("")
                            r.td("")
                            r.td("")
                            r.td("")
        #                                 ws2.cell('F%d'%(rowiterator+6)).value = "DNF"
        #                         rowiterator = rowiterator+1

        #         print(fpart + str(h) + lpart)
        f = open("data.html", "w")
        #         f.write(fpart + script + str(h) + lpart)
        unikoodi = str(h).replace("ä", "&auml;")
        f.write(unikoodi.replace("ö", "&ouml;"))
        f.close()

    def writetoxlsx(self):
        rowiterator = 1
        columniterator = 0
        wb = Workbook()
        # ws = wb.get_active_sheet()
        # ws2 = wb.get_active_sheet() #wb.create_sheet(0) # insert at first position
        ws2 = wb.active
        ws2.title = "Kilpailudatat"
        ws2["A1"].value = "Paikkakunta"
        ws2.cell(1, 2, time.strftime("%d.%m.%Y", time.localtime(time.time())))
        ws2.cell(3, 1, self.__competiontName)
        ws2.cell(4, 1, "Seinäjoki " + time.strftime("%d.%m.%Y", time.localtime(time.time())))
        ws2.cell(5, 5, "Bibnumber")
        ws2.cell(5, 6, "Total")
        ws2.cell(5, 7, "Diff")
        """ws2.cell('A4').value = 'Seinäjoki ' + time.strftime("%d.%m.%Y",time.localtime(time.time()))
            ws2.cell('E5').value = 'Numero'
            ws2.cell('F5').value = 'Total' """
        # ws2.cell('G5').value = 'Time1'
        # ws2.cell('H5').value = 'Time2'
        # ws2.cell('I5').value = 'Time3'
        # self.temporary = sorted(self.competitors, key=lambda kilpailija: kilpailija.totaltime)

        # self.temporary = sorted(self.competitors, key=lambda kilpailija: kilpailija.totaltime)
        # self.temporary2 = sorted(self.competitors, key=lambda kilpailija: kilpailija.valiaikamaara, reverse=True)
        self.temporary2 = sorted(self.competitors, key=attrgetter("totaltime"))
        self.temporary = sorted(self.temporary2, key=attrgetter("valiaikamaara"), reverse=True)
        for luokka in self.luokat:
            ws2.cell(row=(rowiterator + 6), column=2, value=luokka)
            rowiterator = rowiterator + 1
            position = 0
            for obj in self.temporary:
                if luokka == obj.kilpasarja:
                    columniterator = 0
                    if obj.totaltime != 9999999999:
                        position = position + 1
                        ws2.cell((rowiterator + 6), 1, position)
                    ws2.cell((rowiterator + 6), 2, obj.etunimi)
                    ws2.cell((rowiterator + 6), 3, obj.sukunimi)
                    ws2.cell((rowiterator + 6), 4, obj.seura)
                    ws2.cell((rowiterator + 6), 5, obj.bibnumber)
                    if obj.lasttime != 0:
                        if position == 1:
                            nr1pos = obj
                        if obj.dnf == False and obj.dns == False and obj.dsq == False:
                            if obj.GetTimeAmount() > 1:
                                ws2.cell(row=6, column=obj.GetTimeAmount() + 7).value = obj.GetTimeAmount()
                                ws2.cell(row=rowiterator + 6, column=6).value = self.ConvertTimeToStringAccurate(
                                    obj.totaltime
                                )
                                if obj.totaltime - nr1pos.totaltime != 0:
                                    ws2.cell(row=rowiterator + 6, column=7).value = self.ConvertTimeToString(
                                        obj.totaltime - nr1pos.totaltime
                                    )
                                for aika in obj.valiajat:
                                    ws2.cell(
                                        row=rowiterator + 6, column=columniterator + 8
                                    ).value = self.ConvertTimeToStringAccurate(aika)
                                    columniterator = columniterator + 1
                                columniterator = 0
                            else:
                                ws2.cell((rowiterator + 6), 6, self.ConvertTimeToStringAccurate(obj.totaltime))
                                if obj.totaltime - nr1pos.totaltime != 0:
                                    ws2.cell(
                                        (rowiterator + 6),
                                        7,
                                        self.ConvertTimeToString(obj.totaltime - nr1pos.totaltime),
                                    )
                        else:
                            ws2.cell((rowiterator + 6), 6, obj.getStatus())
                    else:
                        ws2.cell((rowiterator + 6), 6, obj.getStatus())

                    rowiterator = rowiterator + 1

            rowiterator = rowiterator + 1
            
        wb.save(
            "xlsx/"
            + self.__competiontName
            + "_"
            + time.strftime("%d_%m_%y_%H_%M_%S", time.localtime(time.time()))
            + ".xlsx"
        )

    def reallyquit(self):
        if self.__competition_running == False:
            ans = askyesno("Veryfy exit", "Oletko varma, että haluat lopettaa ohjelman.")
            if ans == True:
                # if xlsxsupport == True:
                #    self.writetoxlsx()
                self.writeHTML()
                self.writeOfficialTimes(0, False)
                self.dbase = shelve.open("backlog/backlog")
                for obj in self.competitors:
                    self.dbase[obj.bibnumber] = obj
                self.dbase.close()
                exit(0)

        else:
            self.writeToLog("Ohjelmaa ei voida sammuttaa kilpailun ollessa käynnissä")

    def writetotxt(self):
        f = open(
            self.__competiontName + "_" + time.strftime("%d_%m_%y_%H_%M_%S", time.localtime(time.time())) + ".txt", "a"
        )
        f.write(self.log.get("1.0", END))
        f.close()

    def writeOfficialTimes(self, tyyppi, query):
        if query == False:
            f = open(
                "results/"
                + str(self.__competiontName)
                + "results_"
                + time.strftime("%d_%m_%y_%H_%M_%S", time.localtime(time.time()))
                + ".txt",
                "a",
            )
        if query == True:
            f = asksaveasfilename()
        # contents = str('Tilasto_' + time.strftime("%d_%m_%y_%H_%M_%S",time.localtime(time.time())) + '.txt')

        contents = str(self.__competiontName) + str(
            " " + time.strftime("%y_%m_%d %H.%M.%S", time.localtime(time.time())) + "\n\n"
        )
        self.temporary2 = sorted(self.competitors, key=attrgetter("totaltime"))
        self.temporary = sorted(self.temporary2, key=attrgetter("valiaikamaara"), reverse=True)

        for luokka in self.luokat:
            position = 0
            contents = str(contents + luokka + "\n")
            for obj in self.temporary:
                if luokka == obj.kilpasarja:
                    columniterator = 0
                    if obj.totaltime != 9999999999:
                        position = position + 1
                        contents = str(contents + str(position))
                        contents = str(contents + " ")
                    else:
                        contents = str(contents + " ")
                    if tyyppi == 1:
                        contents += str("(" + obj.bibnumber + ") ")
                    contents = str(contents + obj.etunimi + " ")
                    contents = str(contents + obj.sukunimi + " ")
                    contents = str(contents + obj.seura + " ")
                    #                     f.write(obj.bibnumber+" ")
                    if obj.lasttime != 0:
                        if position == 1:
                            nr1pos = obj
                        if obj.dnf == False:
                            if obj.GetTimeAmount() > 1:
                                contents = str(contents + self.ConvertTimeToString(obj.totaltime) + " ")
                                for aika in obj.valiajat:
                                    contents = str(contents + self.ConvertTimeToString(aika) + " ")
                                    columniterator = columniterator + 1
                                if len(obj.valiajat) == 1:
                                    contents = str(contents + " ")
                                    contents = str(contents + " ")
                                if len(obj.valiajat) == 2:
                                    contents = str(contents + " ")

                                columniterator = 0
                            else:
                                if position == 1:
                                    contents = str(contents + self.ConvertTimeToString(obj.totaltime) + " ")
                                else:
                                    contents = str(
                                        contents
                                        + self.ConvertTimeToString(obj.totaltime)
                                        + " +"
                                        + self.ConvertTimeToString(obj.totaltime - nr1pos.totaltime)
                                        + " "
                                    )
                                    print("+" + self.ConvertTimeToString(obj.totaltime - nr1pos.totaltime))
                                contents = str(contents + " ")
                                contents = str(contents + " ")
                                contents = str(contents + " ")
                        else:
                            contents = str(contents + "DNF")
                    else:
                        if obj.dnf == True:
                            contents = str(contents + "DNF")
                        else:
                            contents = str(contents + " ")
                            contents = str(contents + " ")
                            contents = str(contents + " ")
                            contents = str(contents + " ")

                    contents = str(contents + "\n")
            contents = str(contents + "\n")

        print(contents)
        self.writeToLog(contents)
        f.write(contents)
        contents = "0"
        f.close()


def makemenu2(win):
    top = Menu(win)
    win.config(menu=top)
    file = Menu(top)
    file.add_command(label="TXT")
    file.add_command(label="HTML")
    if xlsxsupport == True:
        file.add_command(label="XLSX")
    top.add_cascade(label="Vie tiedot", menu=file)

    kilpailu = Menu(top)
    kilpailu.add_command(label="Start")
    kilpailu.add_command(label="Stop")
    top.add_cascade(label="Kilpailu", menu=kilpailu)


def main():
    root = Tk()
    if mqttsupport == True:
        root.title("Jtimer 1.0")
    else:
        root.title("Jtimer 1.0, NO MQTT SUPPORT")
    mainform = Kilpailu(root)
    mainform.pack()
    root.protocol("WM_DELETE_WINDOW", mainform.reallyquit)

    def insert(event):
        mainform.setfocus()

    def getname(event):
        mainform.getname()

    def valitulostus(event):
        mainform.addline()

    Label(root, text="Yhteislähtöajastus 1.0, viidakkovekara@gmail.com").pack()

    root.bind("<Key-Return>", insert)
    root.bind("l", valitulostus)

    for i in range(10):
        root.bind(str(i), getname)
    root.bind("<Key-BackSpace>", getname)
    root.mainloop()


if __name__ == "__main__":
    if mqttsupport == True:
        portopen = False
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.settimeout(2)
        result = sock.connect_ex(("127.0.0.1", 1883))
        if result == 0:
            portopen = True
        else:
            portopen = False

        if portopen == True:
            mqttc = paho.Client()
            mqttc.on_connect = on_connect
            mqttc.on_message = on_message
            # mqttc.on_log = on_log

            awshost = "127.0.0.1"
            awsport = 1883
            clientId = "clientId"
            thingName = "thingName"
            caPath = "cert/rootCA.pem"
            certPath = "cert/certificate.pem.crt"
            keyPath = "cert/private.pem.key"

            # mqttc.tls_set(caPath, certfile=certPath, keyfile=keyPath, cert_reqs=ssl.CERT_REQUIRED, tls_version=ssl.PROTOCOL_SSLv23, ciphers=None)

            mqttc.connect(awshost, awsport, keepalive=60)

            mqttc.loop_start()
        else:
            mqttsupport = False

    if os.path.exists("results") == False:
        os.mkdir("results")
    if os.path.exists("backlog") == False:
        os.mkdir("backlog")
    if os.path.exists("xlsx") == False:
        os.mkdir("xlsx")
    main()
